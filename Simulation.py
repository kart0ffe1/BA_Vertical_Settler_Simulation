# -*- coding: utf-8 -*-[]
"""
Created on Mon Apr 27 04:02:34 2026

@author: kaku03
"""

import numpy as np
from matplotlib import pyplot as plt
import helper_functions as hlp
from scipy import optimize as optimize
import math
import os
from openpyxl import Workbook

class Batch_Simulation:
    def __init__(self, Exp, Sed_Model):
        self.Exp = Exp
        self.v_0 = Sed_Model.v_0
        self.phi_0 = Sed_Model.phi_0
        
        self.delta_t = 0.5
        self.tol = None
        self.time = None
        self.h_c = None
        self.h_s = None
        self.h_p = None
        self.t_f = None      
        self.psi_i = None
        self.V = None
        self.t_i = None
        self.tau_0 = None
        self.delta_h_i = None
    
    def calc_psi_i(self,t_i):        
        if self.tau_0 is None:
            raise ValueError('tau_0 has not been calculated')
        
        psi_i = (
            self.Exp.Epsilon_p*(2*self.Exp.H_0*(1 - self.Exp.Epsilon_0)
            - self.v_0*t_i)/((1 - self.Exp.Epsilon_p)*(3*self.tau_0 + t_i))    
        )
        return psi_i
          
    
    def calc_V(self,t_i, psi_i):
        if psi_i is None: 
            raise ValueError('psi_i has not been calculated')
        V = (
             self.Exp.H_0/t_i - self.v_0/2 - 
             (1 - self.Exp.Epsilon_p)*psi_i/(2*self.Exp.Epsilon_p)   
        )
        return V
  
    
    def calc_delta_h_i(self):
        if self.v_0 is None:
            raise ValueError('v_0 has not been calculated')
        if self.psi_i is None:
            raise ValueError('psi_i has not been calculated')
        if self.t_i is None:
            raise ValueError('t_i has not been calculated')
        
        delta_h_i = (
            (1 - self.Exp.Epsilon_0)/(1 - self.Exp.Epsilon_p)*self.Exp.H_0
            - self.v_0*self.t_i/(2*(1 - self.Exp.Epsilon_p))      
            - self.psi_i*self.t_i/(2*self.Exp.Epsilon_p)
        )
        return delta_h_i
            
            
            
    def calc_tau_0_exp(self):
        def calc_psi_i():
            max_incl = abs((self.Exp.H_c[2] - self.Exp.H_c[0])/(self.Exp.H_c_time[2]- self.Exp.H_c_time[0]))
            for i in range(3, len(self.Exp.H_c)):
                incl = abs((self.Exp.H_c[i] - self.Exp.H_c[i-2])/(self.Exp.H_c_time[i] -  self.Exp.H_c_time[i-2]))
                if incl > max_incl:
                    max_incl = incl
                    
            self.psi_i = max_incl    
            
        def calc_t_i():
            if self.v_0 is None:
                raise ValueError('v_0 has not been calculated')
            if self.psi_i is None:
                raise ValueError('psi_i has not been calculated') 
            
            
            def f(x):               
                f = (
                    (1 - self.Exp.Epsilon_0)/(1 - self.Exp.Epsilon_p)*self.Exp.H_0 
                    - self.Exp.Epsilon_p*self.v_0*x/(2*(1 - self.Exp.Epsilon_p))
                    - self.psi_i*x/2 - self.Exp.H_0 + self.calc_V(x, self.psi_i)*x
                    + self.psi_i*x/np.log(1 - self.psi_i/self.calc_V(x, self.psi_i))
                )
                return f
            
            def cons_log(x):
                return 1 - self.psi_i/self.calc_V(x, self.psi_i)
            
            t_i_intervals = []
            in_interval = False
            lb = None
            ub = None
            for i in range (1, int(self.Exp.H_c_time[-1])):
                if not in_interval:
                    if cons_log(i) > 0:
                        lb = i
                        in_interval = True
                if in_interval:
                    if cons_log(i) <= 0:
                        ub = i - 1
                        in_interval = False
                        t_i_intervals.append([lb, ub])
            if in_interval:
                ub = int(self.Exp.H_c_time[-1])
                t_i_intervals.append([lb, ub])
                        
            
            t_i = []
            res = []
            
            for bounds in t_i_intervals:
                try:
                    t_i.append(optimize.brentq(f, bounds[0], bounds[1]))
                    res.append(f(t_i[-1]))
                except:
                    pass
            
            if not t_i:
                raise ValueError('t_i could not be determined')
            
            self.t_i = t_i[res.index(min(res))]
            self.V = self.calc_V(self.t_i, self.psi_i)
            
        def calc_tau_0():
            if self.v_0 is None: 
                raise ValueError('v_0 has not been calculated')
            if self.psi_i is None:
                raise ValueError('psi_i has not been calculated')
            if self.t_i is None:
                raise ValueError('t_i has not been calculated')
                
            self.tau_0 = (2*self.Exp.H_0*self.Exp.Epsilon_p*(1 - self.Exp.Epsilon_0) - (self.v_0*self.Exp.Epsilon_p + self.psi_i*(1 - self.Exp.Epsilon_p))*self.t_i)/(3*self.psi_i*(1 - self.Exp.Epsilon_p))
            
            print(f'tau_0 determined to be {self.tau_0}s')
            
        calc_psi_i()
        calc_t_i()
        calc_tau_0()
        self.delta_h_i = self.calc_delta_h_i()
        
    def calc_tau_0_phys(self):            
        def calc_tau_0():
            f = math.pi/6*self.phi_0**3*self.Exp.Delta_rho*self.Exp.G
            r = self.phi_0**2*(self.Exp.Delta_rho*self.Exp.G/(12*self.Exp.sigma))**0.5
            delta_r = 0.267*(math.pi*r**4*self.Exp.A_m**2/(6*self.Exp.sigma*f))**(1/7)
            self.tau_0 = (3*math.pi*self.Exp.Mu_c*r**4)/(4*f*delta_r**2)
            
            print(f'Tau determined to be {self.tau_0}s')
            
        def calc_t_i():                       
            def f(x):               
                f = (
                    (1 - self.Exp.Epsilon_0)/(1 - self.Exp.Epsilon_p)*self.Exp.H_0 
                    - self.Exp.Epsilon_p*self.v_0*x/(2*(1 - self.Exp.Epsilon_p))
                    - self.calc_psi_i(x)*x/2 - self.Exp.H_0 + self.calc_V(x, self.calc_psi_i(x))*x
                    + self.calc_psi_i(x)*x/np.log(1 - self.calc_psi_i(x)/self.calc_V(x, self.calc_psi_i(x)))
                )
                return f
            
            
            
            def cons_log(x):
                return 1 - self.calc_psi_i(x)/self.calc_V(x,  self.calc_psi_i(x))
            
            t_i_intervals = []
            in_interval = False
            lb = None
            ub = None
            for i in range (1, int(self.Exp.H_c_time[-1])):
                if not in_interval:
                    if cons_log(i) > 0:
                        lb = i
                        in_interval = True
                if in_interval:
                    if cons_log(i) <= 0:
                        ub = i - 1
                        in_interval = False
                        t_i_intervals.append([lb, ub])
            if in_interval:
                ub = int(self.Exp.H_c_time[-1])
                t_i_intervals.append([lb, ub])
                        
            
            t_i = []
            res = []
            
            for bounds in t_i_intervals:
                try:
                    t_i.append(optimize.brentq(f, bounds[0], bounds[1]))
                    res.append(f(t_i[-1]))
                except:
                    pass
            
            if not t_i:
                raise ValueError('t_i could not be determined')
            
            self.t_i = t_i[res.index(min(res))]
            self.psi_i = self.calc_psi_i(self.t_i)
            self.V = self.calc_V(self.t_i, self.psi_i)
            
            
        calc_tau_0()
        calc_t_i()
        self.delta_h_i = self.calc_delta_h_i()
        
    def set_parameters(self, v_0, phi_0, tau_0):
        self.v_0 = v_0
        self.phi_0 = phi_0
        self.tau_0 = tau_0       
        
        def t_i_iteration(t_i):
            t_i_iteration = (
                (1 - self.Exp.Epsilon_0)/(1 - self.Exp.Epsilon_p)
                *self.Exp.H_0- self.Exp.Epsilon_p*v_0*t_i
                /(2*(1 - self.Exp.Epsilon_p)) - self.calc_psi_i(t_i)*t_i/2
                - self.Exp.H_0 + self.calc_V(t_i, self.calc_psi_i(t_i))*t_i + self.calc_psi_i(t_i)*t_i
                /np.log(1 - self.calc_psi_i(t_i)/self.calc_V(t_i, self.calc_psi_i(t_i)))
            )
            return t_i_iteration
        
        def cons_log(x):
            return 1 - self.calc_psi_i(x)/self.calc_V(x, self.calc_psi_i(x))
        
        t_i_intervals = []
        in_interval = False
        lb = None
        ub = None
        for i in range (1, int(self.Exp.H_c_time[-1])):
            if not in_interval:
                if cons_log(i) > 0:
                    lb = i
                    in_interval = True
            if in_interval:
                if cons_log(i) <= 0:
                    ub = i - 1
                    in_interval = False
                    t_i_intervals.append([lb, ub])
        if in_interval:
            ub = int(self.Exp.H_c_time[-1])
            t_i_intervals.append([lb, ub])
        
        t_i = []
        res = []
        
        for bounds in t_i_intervals:
            try:
                t_i.append(optimize.brentq(t_i_iteration, bounds[0], bounds[1]))
                res.append(t_i_iteration(t_i[-1]))
            except:
                pass
        
        if not t_i:
            raise ValueError('t_i could not be determined')
        
        self.t_i = t_i[res.index(min(res))]
        self.psi_i = self.calc_psi_i(self.t_i)
        self.V = self.calc_V(self.t_i, self.psi_i)
        self.delta_h_i = self.calc_delta_h_i()
        
    def calc_tau_0_num(self):
        def tau_0_iteration(tau_0):                       
            self.tau_0 = tau_0
            def f(x):               
                f = (
                    (1 - self.Exp.Epsilon_0)/(1 - self.Exp.Epsilon_p)*self.Exp.H_0 
                    - self.Exp.Epsilon_p*self.v_0*x/(2*(1 - self.Exp.Epsilon_p))
                    - self.calc_psi_i(x)*x/2 - self.Exp.H_0 + self.calc_V(x, self.calc_psi_i(x))*x
                    + self.calc_psi_i(x)*x/np.log(1 - self.calc_psi_i(x)/self.calc_V(x, self.calc_psi_i(x)))
                )
                return f
            
            
            
            def cons_log(x):
                return 1 - self.calc_psi_i(x)/self.calc_V(x, self.calc_psi_i(x))
            
            t_i_intervals = []
            in_interval = False
            lb = None
            ub = None
            for i in range (1, int(self.Exp.H_c_time[-1])):
                if not in_interval:
                    if cons_log(i) > 0:
                        lb = i
                        in_interval = True
                if in_interval:
                    if cons_log(i) <= 0:
                        ub = i - 1
                        in_interval = False
                        t_i_intervals.append([lb, ub])
            if in_interval:
                ub = int(self.Exp.H_c_time[-1])
                t_i_intervals.append([lb, ub])
                        
            
            t_i = []
            res = []
            
            for bounds in t_i_intervals:
                try:
                    t_i.append(optimize.brentq(f, bounds[0], bounds[1]))
                    res.append(f(t_i[-1]))
                except:
                    pass
            
            if not t_i:
                raise ValueError('t_i could not be determined')
            
            self.t_i = t_i[res.index(min(res))]
            self.psi_i = self.calc_psi_i(self.t_i)
            self.V = self.calc_V(self.t_i, self.psi_i)
            self.delta_h_i = self.calc_delta_h_i()
            
            self.simulate()
            #self.plot_Results()
            return self.err()
            
            
        best_res = []
        first_found = False
        for i in range(1, 151):
            if not first_found:
                try:
                    err = tau_0_iteration(i)
                    best_res = [i, err]
                    first_found = True
                except:
                    pass
            else:
                try:
                    if tau_0_iteration(i) < best_res[1]:
                        err = tau_0_iteration(i)
                        best_res = [i, err]
                except:
                    pass
            
        
        self.tau_0 = best_res[0]
        print(f'tau 0 found with {self.tau_0}s with an error of {best_res[1]}')
        tau_0_iteration(self.tau_0)
        
    
    
    def simulate(self):
        if self.v_0 is None:
            raise ValueError('Parameters first need to be calculated')
        if self.t_i is None:
            raise ValueError('No t_i determined')
        
        self.time = []
        self.h_c = []
        self.h_s = []
        self.h_p = []
        self.h_c.append(self.Exp.H_0)
        self.h_s.append(0)
        self.h_p.append(self.Exp.H_0)
        t = self.delta_t
        self.time.append(0) 
        self.tol = 30*self.phi_0
        #self.tol = 1e-3
        
        while t < self.t_i:
            self.time.append(t)
            self.h_c.append(
                (self.Exp.H_0 - self.V*t + 
                (self.V*self.delta_h_i)/(self.psi_i)
                *(1 - np.exp(-self.psi_i*t/
                (self.delta_h_i))))
            )
            self.h_s.append(
                (self.v_0*t - (self.v_0 - 
                (1 - self.Exp.Epsilon_p)/(self.Exp.Epsilon_p)*
                self.psi_i)*t**2/(2*self.t_i))
            )
            self.h_p.append(self.Exp.H_0 - self.V*t)
            t += self.delta_t
            
        while self.h_c[-1] - self.h_s[-1] > self.tol:
            self.time.append(t)
            self.h_c.append(
                ((1 - self.Exp.Epsilon_0)*self.Exp.H_0 
                + self.Exp.Epsilon_p*self.delta_h_i
                *np.exp(-self.psi_i*(t-self.t_i)/
                (self.Exp.Epsilon_p*self.delta_h_i)))
            )
            self.h_s.append(
                ((1 - self.Exp.Epsilon_0)*self.Exp.H_0 
                - (1 - self.Exp.Epsilon_p)*self.delta_h_i
                *np.exp(-self.psi_i*(t-self.t_i)/
                (self.Exp.Epsilon_p*self.delta_h_i)))    
            )
            t += self.delta_t
        
        self.t_f = self.time[-1]
            
    
    def err(self):
        if self.h_c is None:
            raise ValueError('Simulation not run yet')
        
        h_c_interp = np.interp(np.array(self.Exp.H_c_time, dtype='float64'), np.array(self.time, dtype='float64'), np.array(self.h_c, dtype='float64'))
        h_s_interp = np.interp(np.array(self.Exp.H_s_time, dtype='float64'), np.array(self.time, dtype='float64'), np.array(self.h_s, dtype='float64'))
        
        err_coal = hlp.nrmse(self.Exp.H_c, h_c_interp, self.Exp.H_0)
        err_sed = hlp.nrmse(self.Exp.H_s, h_s_interp, self.Exp.H_0/(1- self.Exp.Epsilon_0))
        err_t_f = ((self.Exp.T_f - self.t_f)/(2*self.Exp.T_f))**2
        err_tot = err_coal + + err_sed + err_t_f
        
        return err_tot
    
    def plot_Results(self):
        if self.time is None:
            raise ValueError('Simulation not run yet')
            
        plt.figure(num='Simulation')
        plt.scatter(self.Exp.H_c_time, self.Exp.H_c, c='blue', label= 'Coalesence Measured')
        plt.scatter(self.Exp.H_s_time, self.Exp.H_s, c='red', label= 'Sedimentation Measured')
        plt.plot(self.time, self.h_c, c='blue', label='Coalesence')
        plt.plot(self.time, self.h_s, c='red', label='Sediemntation')
        plt.plot(self.time[:len(self.h_p)], self.h_p, c='green', label='Densed Packed Zone')
        plt.vlines(self.t_i, ymin= self.Exp.H_s[0], ymax=self.Exp.H_c[0], linestyles='dashed', colors='black', label= 't_i')
        plt.legend()
        plt.xlabel('Time [s]')
        plt.ylabel('Height [m]')
        plt.show()
            
    def exp_Params(self, path='Export\Simulation_Parameters.xlsx'):
        wb = Workbook()
        ws1 = wb.active
        ws1.titel = "Parameters"
        
        ws1.cell(row=1, column=1, value='tau_0[s]')
        if self.tau_0 is not None:
            ws1.cell(row=2, column=1, value=self.tau_0)
        else: 
            ws1.cell(row=2, column=1, value='-')
            
        ws1.cell(row=1, column=2, value='phi_0[mm]')
        if self.phi_0 is not None:
            ws1.cell(row=2, column=2, value=self.phi_0*1e3)
        else: 
            ws1.cell(row=2, column=2, value='-') 
            
        ws1.cell(row=1, column=3, value='v_0[mm/s]')
        if self.v_0 is not None:
            ws1.cell(row=2, column=3, value=self.v_0*1e3)
        else: 
            ws1.cell(row=2, column=3, value='-') 
            
        ws1.cell(row=1, column=4, value='psi_i[mm/s]')
        if self.psi_i is not None:
            ws1.cell(row=2, column=4, value=self.psi_i*1e3)
        else: 
            ws1.cell(row=2, column=4, value='-')
            
        ws1.cell(row=1, column=5, value='delta_h_i[mm]')
        if self.delta_h_i is not None:
            ws1.cell(row=2, column=5, value=self.delta_h_i*1e3)
        else: 
            ws1.cell(row=2, column=5, value='-')
            
        ws1.cell(row=1, column=6, value='t_i[s]')
        if self.t_i is not None:
            ws1.cell(row=2, column=6, value=self.t_i)
        else: 
            ws1.cell(row=2, column=6, value='-')
            
        ws1.cell(row=1, column=7, value='t_f[s]')
        if self.t_f is not None:
            ws1.cell(row=2, column=7, value=self.t_f)
        else: 
            ws1.cell(row=2, column=7, value='-')
            
        ws1.cell(row=1, column=8, value='V[mm/s]')
        if self.V is not None:
            ws1.cell(row=2, column=8, value=self.V*1e3)
        else: 
            ws1.cell(row=2, column=8, value='-')
            
        ws1.cell(row=1, column=9, value='Err')
        try:
            ws1.cell(row=2, column=9, value=self.err())
        except:
            ws1.cell(row=2, column=9, value='-')
            
        folder = os.path.dirname(path)
        
        if folder:
            if not os.path.exists(folder):
                os.makedirs(folder)
                
        wb.save(path)
        print(f'Parameters have been succesfully Exported to {path}')
        
        
        